"""
gerar_docs_qa.py
Converte planilhas de QA (.xlsx) em documentação Markdown formatada.

Autor: Ian Santos
Dependências: pip install openpyxl

Uso:
    python gerar_docs_qa.py --arquivo "Casos de Testes.xlsx" --projeto detailer-app
    python gerar_docs_qa.py --arquivo "Casos de Testes.xlsx" --projeto detailer-app --tipo bugs
"""

import openpyxl
import argparse
import os
import sys
from datetime import datetime


def ler_planilha(caminho):
    """Lê uma planilha .xlsx e retorna lista de dicionários (uma por linha)."""
    if not os.path.exists(caminho):
        print(f"[ERRO] Arquivo não encontrado: {caminho}")
        sys.exit(1)

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    rows = []

    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        dados = dict(zip(headers, row))
        # Ignorar linhas completamente vazias
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(dados)

    print(f"[INFO] {len(rows)} registros lidos de '{caminho}'")
    return rows


def gerar_casos_de_teste(rows, saida):
    """Gera documento Markdown de casos de teste a partir dos dados da planilha."""
    linhas = ["# Casos de Teste\n\n"]
    aprovados, reprovados, bloqueados, nao_executados = 0, 0, 0, 0
    avisos = []

    for i, row in enumerate(rows, 1):
        ct_id = row.get("ID", f"CT-{i:03d}")
        titulo = row.get("Título", row.get("Titulo", "Sem título"))
        modulo = row.get("Módulo", row.get("Modulo", "—"))
        pre_condicao = row.get("Pré-condição", row.get("Pre-condição", "—"))
        passos = row.get("Passos", row.get("Steps", "—"))
        resultado_esperado = row.get("Resultado Esperado", "—")
        resultado_obtido = row.get("Resultado Obtido", "—")
        status_raw = str(row.get("Status", "")).strip().lower()
        tipo = row.get("Tipo", row.get("Tipo de Teste", "Funcional"))

        # Classificar status
        if any(s in status_raw for s in ["passou", "pass", "ok", "aprovado", "sucesso"]):
            aprovados += 1
            status_icon = "✅ Passou"
        elif any(s in status_raw for s in ["falhou", "fail", "erro", "reprovado", "bug"]):
            reprovados += 1
            status_icon = "❌ Falhou"
        elif any(s in status_raw for s in ["bloqueado", "blocked", "impedido"]):
            bloqueados += 1
            status_icon = "⚠️ Bloqueado"
        else:
            nao_executados += 1
            status_icon = "⬜ Não executado"

        # Avisos de qualidade
        if not resultado_esperado or resultado_esperado == "—":
            avisos.append(f"  ⚠ {ct_id} — sem Resultado Esperado")
        if not passos or passos == "—":
            avisos.append(f"  ⚠ {ct_id} — sem Passos de execução")

        # Montar caso de teste
        linhas.append(f"## {ct_id} — {titulo}\n\n")
        linhas.append(f"| Campo | Valor |\n|-------|-------|\n")
        linhas.append(f"| **Módulo** | {modulo} |\n")
        linhas.append(f"| **Tipo** | {tipo} |\n")
        linhas.append(f"| **Pré-condição** | {pre_condicao} |\n")
        linhas.append(f"| **Status** | {status_icon} |\n\n")

        if passos and passos != "—":
            linhas.append(f"### Passos\n{passos}\n\n")

        linhas.append(f"**Resultado esperado:** {resultado_esperado}\n\n")

        if resultado_obtido and resultado_obtido != "—":
            linhas.append(f"**Resultado obtido:** {resultado_obtido}\n\n")

        linhas.append("---\n\n")

    # Métricas
    total = aprovados + reprovados + bloqueados + nao_executados
    taxa = (aprovados / total * 100) if total > 0 else 0

    linhas.append(f"## 📊 Métricas\n\n")
    linhas.append(f"| Métrica | Valor |\n|---------|-------|\n")
    linhas.append(f"| Total de Casos | {total} |\n")
    linhas.append(f"| ✅ Aprovados | {aprovados} |\n")
    linhas.append(f"| ❌ Reprovados | {reprovados} |\n")
    linhas.append(f"| ⚠️ Bloqueados | {bloqueados} |\n")
    linhas.append(f"| ⬜ Não Executados | {nao_executados} |\n")
    linhas.append(f"| **Taxa de Aprovação** | **{taxa:.0f}%** |\n")

    with open(saida, "w", encoding="utf-8") as f:
        f.writelines(linhas)

    print(f"\n[OK] Gerado: {saida}")
    print(f"     {total} casos | ✅ {aprovados} | ❌ {reprovados} | ⚠️ {bloqueados} | ⬜ {nao_executados}")
    print(f"     Taxa de aprovação: {taxa:.0f}%")

    if avisos:
        print(f"\n[AVISOS DE QUALIDADE] {len(avisos)} problemas encontrados:")
        for aviso in avisos:
            print(aviso)


def gerar_bugs(rows, saida):
    """Gera documento Markdown de bugs a partir dos dados da planilha."""
    linhas = ["# Bugs Report\n\n"]
    total_alta = 0
    total_media = 0
    total_baixa = 0

    for i, row in enumerate(rows, 1):
        bug_id = row.get("ID", row.get("Bug ID", f"BUG-{i:03d}"))
        titulo = row.get("Título", row.get("Titulo", row.get("Descrição", "Sem título")))
        modulo = row.get("Módulo", row.get("Modulo", "—"))
        severidade = str(row.get("Severidade", "—")).strip()
        prioridade = str(row.get("Prioridade", "—")).strip()
        status = row.get("Status", "Aberto")
        passos = row.get("Passos", row.get("Passos para Reproduzir", "—"))
        resultado_esperado = row.get("Resultado Esperado", "—")
        resultado_obtido = row.get("Resultado Obtido", "—")

        if "alta" in severidade.lower():
            total_alta += 1
        elif "média" in severidade.lower() or "media" in severidade.lower():
            total_media += 1
        else:
            total_baixa += 1

        linhas.append(f"## {bug_id} — {titulo}\n\n")
        linhas.append(f"| Campo | Valor |\n|-------|-------|\n")
        linhas.append(f"| **Módulo** | {modulo} |\n")
        linhas.append(f"| **Severidade** | {severidade} |\n")
        linhas.append(f"| **Prioridade** | {prioridade} |\n")
        linhas.append(f"| **Status** | {status} |\n\n")

        if passos and passos != "—":
            linhas.append(f"### Passos para Reproduzir\n{passos}\n\n")

        linhas.append(f"**Resultado esperado:** {resultado_esperado}\n\n")

        if resultado_obtido and resultado_obtido != "—":
            linhas.append(f"**Resultado obtido:** {resultado_obtido}\n\n")

        linhas.append("---\n\n")

    # Resumo
    total = total_alta + total_media + total_baixa
    linhas.insert(1, f"| Métrica | Valor |\n|---------|-------|\n"
                     f"| Total de Bugs | {total} |\n"
                     f"| Alta Severidade | {total_alta} |\n"
                     f"| Média Severidade | {total_media} |\n"
                     f"| Baixa / Sem classificação | {total_baixa} |\n\n---\n\n")

    with open(saida, "w", encoding="utf-8") as f:
        f.writelines(linhas)

    print(f"\n[OK] Gerado: {saida}")
    print(f"     {total} bugs | 🔴 {total_alta} alta | 🟡 {total_media} média | 🟢 {total_baixa} baixa")


def main():
    parser = argparse.ArgumentParser(
        description="Converte planilhas QA (.xlsx) em Markdown",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python gerar_docs_qa.py --arquivo planilha.xlsx --projeto detailer-app
  python gerar_docs_qa.py --arquivo bugs.xlsx --projeto barber-os --tipo bugs
        """
    )
    parser.add_argument("--arquivo", required=True, help="Caminho da planilha .xlsx")
    parser.add_argument("--projeto", required=True, help="Nome da pasta do projeto")
    parser.add_argument(
        "--tipo",
        choices=["casos", "bugs"],
        default="casos",
        help="Tipo de documento a gerar (padrão: casos)"
    )
    parser.add_argument(
        "--saida",
        help="Caminho de saída (padrão: projetos/<projeto>/casos-de-teste.md)"
    )

    args = parser.parse_args()

    rows = ler_planilha(args.arquivo)

    if args.saida:
        saida = args.saida
    else:
        nome_arquivo = "casos-de-teste.md" if args.tipo == "casos" else "bugs-reportados.md"
        saida = os.path.join("projetos", args.projeto, nome_arquivo)

    os.makedirs(os.path.dirname(saida), exist_ok=True)

    if args.tipo == "casos":
        gerar_casos_de_teste(rows, saida)
    else:
        gerar_bugs(rows, saida)

    print(f"\n{'='*50}")
    print(f"Documento gerado com sucesso!")
    print(f"Arquivo: {saida}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
